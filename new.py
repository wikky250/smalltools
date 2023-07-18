import os
from shutil import copyfile
from openpyxl import load_workbook
import handleCNY
import random
from openpyxl import Workbook
from openpyxl.cell import MergedCell
import operator

# wb = load_workbook("D:/报价单1.xlsx")
# ws = wb.active
# # 需要合并的左上方和右下方单元格坐标
# sheet = wb.active
# shoulddeleteup = 0
# shoulddeletedown = 0
# for i in range(0,10):
#     cell = sheet.cell(6+i, 3)
#     if cell.coordinate in ws.merged_cells:
#         value = sheet.cell(6+i, 2).value
#         if operator.contains(str(value),"合计"):
#             shoulddeletedown = 6+i
#             break
#     else:
#         ws.merge_cells(range_string='H'+str(6+i)+':'+'I'+str(6+i))
#         shoulddeleteup = 6+i

# # ws.move_range('B11:B12', rows=shoulddeleteup-shoulddeletedown, cols=0, translate=True)

# for i in range(shoulddeletedown-1,shoulddeleteup-1,-1):
#     ws.delete_rows(i)
# ws.merge_cells(start_row=shoulddeleteup, start_column=2, end_row=shoulddeleteup, end_column=3)
# ws.unmerge_cells(start_row=shoulddeleteup, start_column=4, end_row=shoulddeleteup, end_column=5)
# ws.merge_cells(start_row=shoulddeleteup, start_column=4, end_row=shoulddeleteup, end_column=7)
# ws.unmerge_cells(start_row=shoulddeleteup+1, start_column=2, end_row=shoulddeleteup+1, end_column=3)
# ws.unmerge_cells(start_row=shoulddeleteup+1, start_column=4, end_row=shoulddeleteup+1, end_column=7)
# ws.unmerge_cells(start_row=shoulddeleteup+1, start_column=8, end_row=shoulddeleteup+1, end_column=9)
# ws.merge_cells(start_row=shoulddeleteup+1, start_column=2, end_row=shoulddeleteup+1, end_column=9)
# wb.save("D:/te.xlsx")


globalfile=""
source_file1="D:/smalltools/test/报价单1.xlsx"
source_file2="D:/smalltools/test/报价单2.xlsx"
def recursive_listdir(path):
    change = handleCNY.ChangeCNY()
    files = os.listdir(path)
    if(len(files)!=0):
        for file in files:
            file_path = os.path.join(path, file)

            if os.path.isfile(file_path):
                print()

            elif os.path.isdir(file_path):
                recursive_listdir(file_path)
    else:
        name = path.split(' ')
        destination_file1 = path+"/报价单1-"+name[1] +".xlsx"
        destination_file2 = path+"/报价单2-"+name[1] +".xlsx"
        copyfile(source_file1, destination_file1)
        copyfile(source_file2, destination_file2)
        wb = load_workbook(destination_file1)
        sheet = wb.active
        sheet['B3'] = 'Devansh Sharma'
        sheet['C4'] = 'hello world'
        value = round(random.random()*100000,2)
        sheet['D7'] = str(value)+"（"+change.small_to_big(str(value))+"）"
        
        wb.save(destination_file1)
        print(path)

recursive_listdir(r'D:/smalltools/test')
